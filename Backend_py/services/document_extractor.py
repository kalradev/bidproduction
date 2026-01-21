import io
import logging
import pdfplumber
import docx
import mammoth
import openpyxl
import pytesseract
from PIL import Image
from typing import Dict, Any, Optional
import re

logger = logging.getLogger(__name__)

async def extract_text(buffer: bytes, mimetype: str, filename: str) -> Dict[str, Any]:
    result = {}
    
    if mimetype == 'application/pdf':
        result = await extract_from_pdf(buffer)
    elif mimetype == 'application/vnd.openxmlformats-officedocument.wordprocessingml.document':
        result = await extract_from_docx(buffer)
    elif mimetype == 'application/msword':
        # Legacy DOC is tricky in Python without external tools like antiword or libreoffice
        # For now, we'll try mammoth or a fallback
        result = await extract_from_docx(buffer) # Fallback attempt
    elif mimetype in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'application/vnd.ms-excel', 'application/excel']:
        result = await extract_from_excel(buffer)
    elif mimetype.startswith('image/'):
        result = await extract_from_image(buffer, filename)
    else:
        raise Exception(f"Unsupported file type: {mimetype}")

    cleaned_text = clean_text(result.get("text", ""))
    
    # Estimate pages
    estimated_pages = result.get("metadata", {}).get("pages", 1)
    if not estimated_pages or estimated_pages == 1:
        word_count = count_words(cleaned_text)
        estimated_pages = max(1, word_count // 500)
        if "metadata" not in result:
            result["metadata"] = {}
        result["metadata"]["pages"] = estimated_pages

    return {
        "text": cleaned_text,
        "originalText": result.get("text", ""),
        "html": result.get("html"),
        "metadata": result.get("metadata", {}),
        "wordCount": count_words(cleaned_text)
    }

async def extract_from_pdf(buffer: bytes) -> Dict[str, Any]:
    try:
        text = ""
        page_count = 0
        with pdfplumber.open(io.BytesIO(buffer)) as pdf:
            page_count = len(pdf.pages)
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
        
        return {
            "text": text,
            "metadata": {"pages": page_count}
        }
    except Exception as e:
        logger.error(f"PDF extraction failed: {str(e)}")
        raise Exception(f"PDF extraction failed: {str(e)}")

async def extract_from_docx(buffer: bytes) -> Dict[str, Any]:
    try:
        # Get raw text using python-docx
        doc = docx.Document(io.BytesIO(buffer))
        text = "\n".join([para.text for para in doc.paragraphs])
        
        # Get HTML using mammoth
        html_result = mammoth.convert_to_html(io.BytesIO(buffer))
        
        return {
            "text": text,
            "html": html_result.value,
            "metadata": {"messages": [m.message for m in html_result.messages]}
        }
    except Exception as e:
        logger.error(f"DOCX extraction failed: {str(e)}")
        raise Exception(f"DOCX extraction failed: {str(e)}")

async def extract_from_excel(buffer: bytes) -> Dict[str, Any]:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(buffer), data_only=True)
        all_text = []
        sheet_data = []
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                filtered_row = [str(cell) for cell in row if cell is not None and str(cell).strip() != ""]
                if filtered_row:
                    rows.append(" | ".join(filtered_row))
            
            sheet_text = "\n".join(rows)
            all_text.append(f"=== Sheet: {sheet_name} ===\n{sheet_text}")
            sheet_data.append({
                "name": sheet_name,
                "rows": len(rows),
                "text": sheet_text
            })
            
        return {
            "text": "\n\n".join(all_text),
            "metadata": {
                "sheets": wb.sheetnames,
                "sheetCount": len(wb.sheetnames),
                "sheetData": sheet_data
            }
        }
    except Exception as e:
        logger.error(f"Excel extraction failed: {str(e)}")
        raise Exception(f"Excel extraction failed: {str(e)}")

async def extract_from_image(buffer: bytes, filename: str) -> Dict[str, Any]:
    try:
        img = Image.open(io.BytesIO(buffer))
        text = pytesseract.image_to_string(img)
        
        return {
            "text": text,
            "metadata": {
                "format": "image (OCR)",
                "filename": filename
            }
        }
    except Exception as e:
        logger.error(f"Image OCR failed: {str(e)}")
        raise Exception(f"Image OCR failed: {str(e)}")

def clean_text(text: str) -> str:
    if not text:
        return ""
    text = text.replace('\r\n', '\n')
    text = re.sub(r'\n{3,}', '\n\n', text)
    text = re.sub(r'[ \t]+', ' ', text)
    return text.strip()

def count_words(text: str) -> int:
    if not text:
        return 0
    return len(re.findall(r'\w+', text))
